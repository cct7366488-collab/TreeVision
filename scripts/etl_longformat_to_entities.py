# -*- coding: utf-8 -*-
"""
ETL：土肉桂試驗「長格式」XLSX → TreeVision v0.2 正規化實體記錄。

將 `土肉桂試驗_長格式.xlsx`（分析就緒長表）轉成四層架構的各實體記錄
（treatment / site / plot / tree / tree_measurement / campaign），逐筆對
`schemas/*.schema.json`（Draft 2020-12）驗證，輸出每實體一份 CSV 與一份
資料品質（DQ）報告。

設計原則
--------
- **不硬編客戶路徑**：輸入以 --longformat 或環境變數 TREEVISION_LONGFORMAT 指定。
- **輸出留本機**：預設寫 outputs/entities/（已被 .gitignore 排除），因衍生自
  客戶試驗資料（含樣木資訊），依 Cloud.md 不上 GitHub。
- **可重現**：純轉換，無外部狀態；DB loader 之後再以 adapter 接上。

用法
----
    python scripts/etl_longformat_to_entities.py \
        --longformat "<path>/土肉桂試驗_長格式.xlsx" \
        --outdir outputs/entities
"""
import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict

from openpyxl import load_workbook
from jsonschema import Draft202012Validator

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
SCHEMA_DIR = os.path.join(REPO, "schemas")

# 場域靜態事實（已見於公開 ADR-0002，可安全內嵌）
SITE_FACTS = {
    "115-12-1": dict(name="大安溪115林班假12號土肉桂試驗區", forest_district="大安溪事業區",
                     compartment="115", sub_lot_no="12", planted_date="2010-10", age_class="mature"),
    "117-28-1": dict(name="大安溪117林班假28號土肉桂幼齡林", forest_district="大安溪事業區",
                     compartment="117", sub_lot_no="28", planted_date="2025-04", age_class="juvenile"),
    "8-1-1":    dict(name="八仙山20林班假1號土肉桂幼齡林", forest_district="八仙山事業區",
                     compartment="20", sub_lot_no="1", planted_date="2024-03", age_class="juvenile"),
    "8-1-10":   dict(name="八仙山20林班假1號-10土肉桂幼齡林", forest_district="八仙山事業區",
                     compartment="20", sub_lot_no="1", planted_date="2024-03", age_class="juvenile"),
}
OWNER, REGION, TOWNSHIP = "林業及自然保育署臺中分署", "臺中市", "和平區"


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return None


def to_date(v):
    if v is None or v == "":
        return None
    s = str(v)
    return s[:10]  # datetime 或 'YYYY-MM-DD...' → 'YYYY-MM-DD'


def site_dash(forest_label):
    m = re.search(r"\(([0-9]+-[0-9]+-[0-9]+)\)", str(forest_label or ""))
    return m.group(1) if m else None


def load_validators():
    out = {}
    for name in ("treatment", "site_registry", "plot", "tree_registry",
                 "tree_measurement", "campaign"):
        path = os.path.join(SCHEMA_DIR, name + ".schema.json")
        out[name] = Draft202012Validator(json.load(open(path, encoding="utf-8")))
    return out


def transform(longformat_path):
    wb = load_workbook(longformat_path, data_only=True, read_only=True)
    ws = wb["長格式"]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}

    def c(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    treatments, sites, plots, trees, campaigns = (OrderedDict() for _ in range(5))
    measurements = []
    dq = []  # 資料品質旗標列
    tree_last = {}  # tree_id -> (season, status) 取最後

    for rn, r in enumerate(rows, start=2):
        site = site_dash(c(r, "林班地"))
        treat = str(c(r, "處理代碼") or "").strip()
        code = str(c(r, "樣木代碼") or "").strip()
        season = str(c(r, "季別") or "").strip()
        flag = str(c(r, "解析旗標") or "").strip()
        tree_no = num(c(r, "樣木編號"))
        if not (site and treat and code and season):
            dq.append(dict(row=rn, code=code, issue="缺少必要鍵(site/treat/code/season)"))
            continue

        base_tree = re.sub(r"-s[0-9]+$", "", code)
        stem_m = re.search(r"-s([0-9]+)$", code)
        stem_seq = int(stem_m.group(1)) if stem_m else 1

        # DQ：樣木編號為空/0（長格式 leading_blank_treeno）→ 無法入 registry，連量測一併隔離
        if flag == "leading_blank_treeno" or not tree_no:
            dq.append(dict(row=rn, code=code, season=season, issue="leading_blank_treeno(編號空白補0)",
                           dbh=num(c(r, "直徑_cm")), height=num(c(r, "樹高_m"))))
            continue

        # treatment
        if treat not in treatments:
            treatments[treat] = dict(treatment_id=treat, label_zh=str(c(r, "處理中文") or "").strip(),
                                     pruning=num(c(r, "修剪")), fertilizer_g=num(c(r, "施肥g")))
        # site
        if site not in sites:
            f = SITE_FACTS.get(site, {})
            sites[site] = dict(site_id=site, site_code_short=str(c(r, "樣區代碼") or "").strip() or None,
                               plot_type="永久樣區", name=f.get("name", site), owner=OWNER,
                               region=REGION, township=TOWNSHIP, forest_district=f.get("forest_district"),
                               compartment=f.get("compartment"), sub_lot_no=f.get("sub_lot_no"),
                               gps_datum="TWD97", centroid_lat=None, centroid_lon=None,
                               area_ha=num(c(r, "面積ha")), age_stage=str(c(r, "林齡階段") or "").strip() or None,
                               age_class=f.get("age_class"), planted_date=f.get("planted_date"))
        # plot
        plot_id = site + "." + treat
        if plot_id not in plots:
            plots[plot_id] = dict(plot_id=plot_id, site_id=site, treatment_id=treat, _trees=set())
        plots[plot_id]["_trees"].add(base_tree)
        # tree
        if base_tree not in trees:
            trees[base_tree] = dict(tree_id=base_tree, site_id=site, plot_id=plot_id, treatment_id=treat,
                                    tree_no=tree_no, species_zh="土肉桂",
                                    species_sci="Cinnamomum osmophloeum",
                                    is_multistem=False, stem_count=1)
        if stem_seq > 1:
            trees[base_tree]["is_multistem"] = True
            trees[base_tree]["stem_count"] = max(trees[base_tree]["stem_count"], stem_seq)
        # campaign
        cid = site + "_" + season
        if cid not in campaigns:
            campaigns[cid] = dict(campaign_id=cid, site_id=site, season=season,
                                  date_estimated=to_date(c(r, "量測日期_推定")), operator="合作社")
        # measurement
        status = str(c(r, "狀態") or "").strip() or None
        measurements.append(dict(tree_id=base_tree, campaign_id=cid, season=season,
                                 measure_date=to_date(c(r, "量測日期_推定")), stem_seq=stem_seq,
                                 measure_part=str(c(r, "量測部位") or "").strip() or None,
                                 dbh_cm=num(c(r, "直徑_cm")), height_m=num(c(r, "樹高_m")),
                                 crown_w1_m=num(c(r, "冠幅NS_m")), crown_w2_m=num(c(r, "冠幅EW_m")),
                                 volume_m3=num(c(r, "材積_原表")), growth_increment=num(c(r, "生長量_原表")),
                                 status=status, measured_by=None))
        # 追蹤最後狀態（季別字典序對 114Q3<114Q4<115Q1 成立）
        key = (season, status)
        if base_tree not in tree_last or season >= tree_last[base_tree][0]:
            tree_last[base_tree] = key

    # 收尾：plot tree_count、tree.status（取最後季別觀測）
    for p in plots.values():
        p["tree_count"] = len(p.pop("_trees"))
    for tid, t in trees.items():
        last = tree_last.get(tid)
        t["status"] = last[1] if last else None

    return dict(treatment=list(treatments.values()), site_registry=list(sites.values()),
                plot=list(plots.values()), tree_registry=list(trees.values()),
                tree_measurement=measurements, campaign=list(campaigns.values())), dq


def validate(entities, validators):
    errors = {}
    for name, recs in entities.items():
        v = validators[name]
        errs = []
        for i, rec in enumerate(recs):
            clean = {k: val for k, val in rec.items() if val is not None}
            for e in v.iter_errors(clean):
                errs.append((i, e.message))
        errors[name] = errs
    return errors


def write_csv(path, recs):
    if not recs:
        open(path, "w", encoding="utf-8").close()
        return
    cols = list(recs[0].keys())
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for rec in recs:
            w.writerow({k: ("" if rec.get(k) is None else rec.get(k)) for k in cols})


def main():
    ap = argparse.ArgumentParser(description="長格式 XLSX → TreeVision v0.2 實體記錄")
    ap.add_argument("--longformat", default=os.environ.get("TREEVISION_LONGFORMAT"),
                    help="土肉桂試驗_長格式.xlsx 路徑（或設環境變數 TREEVISION_LONGFORMAT）")
    ap.add_argument("--outdir", default=os.path.join(REPO, "outputs", "entities"))
    args = ap.parse_args()
    if not args.longformat or not os.path.exists(args.longformat):
        ap.error("需指定有效的 --longformat 路徑（客戶試驗檔，本機）")

    os.makedirs(args.outdir, exist_ok=True)
    validators = load_validators()
    entities, dq = transform(args.longformat)
    errors = validate(entities, validators)

    print("=== 實體記錄數 / schema 驗證 ===")
    total_err = 0
    for name, recs in entities.items():
        ne = len(errors[name])
        total_err += ne
        write_csv(os.path.join(args.outdir, name + ".csv"), recs)
        tag = "OK " if ne == 0 else "ERR"
        print("{} {:20s} records={:5d} schema_errors={}".format(tag, name, len(recs), ne))
        for i, msg in errors[name][:3]:
            print("      rec#{}: {}".format(i, msg))

    report = dict(longformat=os.path.basename(args.longformat),
                  record_counts={k: len(v) for k, v in entities.items()},
                  schema_errors={k: len(v) for k, v in errors.items()},
                  dq_flag_count=len(dq), dq_flags=dq)
    with open(os.path.join(args.outdir, "dq_report.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print("\n=== 資料品質（DQ）===")
    print("隔離旗標列數:", len(dq), "（多為 leading_blank_treeno；詳見 dq_report.json）")
    print("輸出目錄:", args.outdir)
    print("\nRESULT:", "ALL VALID" if total_err == 0 else "SCHEMA ERRORS={}".format(total_err))
    sys.exit(0 if total_err == 0 else 1)


if __name__ == "__main__":
    main()
